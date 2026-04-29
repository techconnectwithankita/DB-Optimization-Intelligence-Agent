from __future__ import annotations

import json
import os
import csv
from io import BytesIO, StringIO
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .services.agent import ApiDesignerAgent


def load_local_env() -> None:
    env_path = Path(__file__).resolve().parents[1] / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_local_env()


class Requirement(BaseModel):
    id: str
    title: str
    desc: str
    source: str | None = None
    priority: str | None = None
    method: str | None = None
    path: str | None = None
    summary: str | None = None


class GenerateRequest(BaseModel):
    requirement: Requirement
    sources: list[str] = Field(default_factory=list)
    domain: str = "Policy Management"
    style: str = "REST"


class ValidationRequest(BaseModel):
    openapi_yaml: str


class ArtifactRequest(BaseModel):
    artifact_type: str
    design: dict[str, Any]


app = FastAPI(title="API Designer Agent", version="1.0.0")

allowed_origins = [
    origin.strip()
    for origin in os.getenv("ALLOWED_ORIGINS", "http://127.0.0.1:5173,http://localhost:5173").split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

designer = ApiDesignerAgent()
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DIST_DIR = PROJECT_ROOT / "dist"


@app.get("/health")
def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "agent": "API Designer Agent",
        "model": designer.model,
        "mock": designer.mock_mode,
        "time": datetime.now().isoformat(timespec="seconds"),
    }


@app.post("/api/design")
async def generate_design(payload: GenerateRequest) -> dict[str, Any]:
    return await designer.generate_design(payload.model_dump())


@app.post("/api/requirements/upload")
async def upload_requirements(file: UploadFile = File(...)) -> dict[str, Any]:
    name = file.filename or "requirements"
    content = await file.read()
    suffix = Path(name).suffix.lower()

    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(text)))
    elif suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="Excel support is not installed.") from exc

        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    else:
        raise HTTPException(status_code=400, detail="Upload a .xlsx, .xlsm, or .csv file.")

    requirements = [normalize_requirement(row, index + 1, name) for index, row in enumerate(rows)]
    return {
        "filename": name,
        "count": len(requirements),
        "requirements": requirements,
        "raw_text": raw_rows_to_text(rows),
        "summary": summarize_requirements(requirements, name),
    }


def normalize_requirement(row: dict[str, Any], index: int, filename: str) -> dict[str, Any]:
    normalized = {str(key).strip().lower().replace(" ", "_"): value for key, value in row.items() if key}

    req_id = str(normalized.get("id") or normalized.get("requirement_id") or f"UP-{index:03d}")
    title = str(normalized.get("title") or normalized.get("requirement") or normalized.get("name") or f"Uploaded Requirement {index}")
    desc = str(normalized.get("description") or normalized.get("desc") or normalized.get("user_story") or title)
    method = str(normalized.get("method") or infer_method(title)).lower()
    path = str(normalized.get("path") or infer_path(title))

    return {
        "id": req_id,
        "title": title,
        "desc": desc,
        "source": str(normalized.get("source") or f"Uploaded: {filename}"),
        "priority": str(normalized.get("priority") or "Medium").title(),
        "method": method,
        "path": path,
        "summary": str(normalized.get("summary") or title),
    }


def infer_method(title: str) -> str:
    text = title.lower()
    if any(word in text for word in ["fetch", "get", "retrieve", "list", "search"]):
        return "get"
    if any(word in text for word in ["update", "change", "modify"]):
        return "patch"
    if any(word in text for word in ["delete", "remove"]):
        return "delete"
    return "post"


def infer_path(title: str) -> str:
    words = re_words(title)
    if "customer" in words:
        base = "/customers"
    elif "claim" in words:
        base = "/claims"
    else:
        base = "/policies"

    if any(word in words for word in ["validate", "validation"]):
        return f"{base}/validate"
    if any(word in words for word in ["status", "update"]):
        return f"{base}/{{id}}/status"
    if any(word in words for word in ["cancel"]):
        return f"{base}/{{id}}/cancel"
    if any(word in words for word in ["fetch", "get", "retrieve"]):
        return f"{base}/{{id}}"
    return base


def re_words(text: str) -> set[str]:
    import re

    return set(re.findall(r"[a-z0-9]+", text.lower()))


def raw_rows_to_text(rows: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    for index, row in enumerate(rows, start=1):
        lines.append(f"Row {index}")
        for key, value in row.items():
            lines.append(f"  {key}: {value}")
    return "\n".join(lines)


def summarize_requirements(requirements: list[dict[str, Any]], filename: str) -> dict[str, Any]:
    priorities: dict[str, int] = {}
    methods: dict[str, int] = {}
    resources: dict[str, int] = {}

    for req in requirements:
        priorities[req["priority"]] = priorities.get(req["priority"], 0) + 1
        methods[req["method"].upper()] = methods.get(req["method"].upper(), 0) + 1
        resource = "/" + req["path"].strip("/").split("/")[0] if req.get("path") else "/"
        resources[resource] = resources.get(resource, 0) + 1

    return {
        "source": filename,
        "total_requirements": len(requirements),
        "priorities": priorities,
        "methods": methods,
        "resources": resources,
        "narrative": (
            f"Extracted {len(requirements)} functional requirements from {filename}. "
            "The source describes API operations, request/response behavior, priorities, and resource paths "
            "that can be converted into OpenAPI endpoints and design artifacts."
        ),
    }


@app.post("/api/validate")
def validate_spec(payload: ValidationRequest) -> dict[str, Any]:
    try:
        parsed = yaml.safe_load(payload.openapi_yaml)
    except yaml.YAMLError as exc:
        return {
            "valid": False,
            "errors": [f"YAML parse error: {exc}"],
            "warnings": [],
            "summary": "OpenAPI validation failed.",
        }

    errors: list[str] = []
    warnings: list[str] = []

    if not isinstance(parsed, dict):
        errors.append("Document must be a YAML object.")
    else:
        if not str(parsed.get("openapi", "")).startswith("3."):
            errors.append("OpenAPI version must be 3.x.")
        if not parsed.get("info"):
            errors.append("Missing required info object.")
        if not parsed.get("paths"):
            errors.append("Missing required paths object.")
        if "components" not in parsed:
            warnings.append("No reusable components section found.")
        if "servers" not in parsed:
            warnings.append("No servers section found.")

    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "summary": "Validation passed." if not errors else "OpenAPI validation failed.",
    }


@app.post("/api/artifact", response_class=PlainTextResponse)
def artifact(payload: ArtifactRequest) -> PlainTextResponse:
    design = payload.design
    artifact_type = payload.artifact_type.lower()

    if artifact_type == "openapi":
        text = design.get("openapi_yaml", "")
        media_type = "application/yaml"
    elif artifact_type == "summary":
        text = design.get("endpoint_summary_markdown", "")
        media_type = "text/markdown"
    elif artifact_type == "schemas":
        text = json.dumps(design.get("schemas_json", {}), indent=2)
        media_type = "application/json"
    elif artifact_type == "postman":
        text = json.dumps(design.get("postman_collection", {}), indent=2)
        media_type = "application/json"
    elif artifact_type == "sequence":
        text = design.get("sequence_diagram_mermaid", "")
        media_type = "text/plain"
    elif artifact_type == "review":
        text = design.get("design_review_markdown", "")
        media_type = "text/markdown"
    elif artifact_type == "swagger":
        text = design.get("swagger_documentation_markdown", "")
        media_type = "text/markdown"
    elif artifact_type == "devkit":
        text = json.dumps(design.get("development_kit", {}), indent=2)
        media_type = "application/json"
    elif artifact_type == "tests":
        text = json.dumps(design.get("testing_package", {}), indent=2)
        media_type = "application/json"
    elif artifact_type == "deployment":
        text = json.dumps(design.get("deployment_package", {}), indent=2)
        media_type = "application/json"
    elif artifact_type == "gateway":
        text = json.dumps(design.get("gateway_setup", {}), indent=2)
        media_type = "application/json"
    elif artifact_type == "monitoring":
        text = json.dumps(design.get("monitoring_readiness", {}), indent=2)
        media_type = "application/json"
    else:
        raise HTTPException(status_code=400, detail=f"Unknown artifact type: {payload.artifact_type}")

    return PlainTextResponse(text, media_type=media_type)


if DIST_DIR.exists():
    assets_dir = DIST_DIR / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")


@app.get("/")
def serve_frontend():
    index_file = DIST_DIR / "index.html"
    if index_file.exists():
        return FileResponse(index_file)
    return {"message": "API Designer Agent backend is running. Build the React frontend to serve the UI."}


@app.get("/{full_path:path}")
def serve_spa(full_path: str):
    if full_path.startswith("api/") or full_path == "health":
        raise HTTPException(status_code=404, detail="Not found")

    index_file = DIST_DIR / "index.html"
    if index_file.exists():
        return FileResponse(index_file)
    return {"message": "Frontend build not found."}
